
import psycopg2
from openpyxl import load_workbook
from config import DATABASE_CONFIG


# Connect to the database
conn = psycopg2.connect(
    host=DATABASE_CONFIG['host'],
    database=DATABASE_CONFIG['database'],
    user=DATABASE_CONFIG['user'],
    password=DATABASE_CONFIG['password']
)

# Create a cursor
cur = conn.cursor()

# Load the workbook
wb = load_workbook('Fall 23 Die Tourney Data.xlsx')

# Select the active sheet
ws = wb.active

# Iterate through the rows of the sheet
for row in ws.iter_rows(min_row=3):
  date = row[0].value
  player1_name = row[1].value
  player2_name = row[2].value
  team1_score = row[3].value
  player3_name = row[4].value
  player4_name = row[5].value
  team2_score = row[6].value

   # Check if the player name is not empty
  if player1_name:
    # Check if the player already exists in the Player table
    cur.execute("SELECT player_id FROM player WHERE first_name=%s", (player1_name,))
    player1_id = cur.fetchone()
    if player1_id is None:
      # If the player does not exist, insert them into the players table with a unique id
      cur.execute("SELECT nextval('player_id_seq')")
      id = cur.fetchone()[0]
      cur.execute("INSERT INTO player (player_id, first_name) VALUES (%s, %s)", (id, player1_name))
      player1_id = id
    else:
      # If the player already exists, retrieve their player_id
      player1_id = player1_id[0]
    # Print the first_name and id of the player
    print(f'Processing player1 {player1_name} with id {player1_id}')

  
# Check if the player2 first_name is not empty
  if player2_name:
    # Check if the player already exists in the Players table
    cur.execute("SELECT player_id FROM player WHERE first_name=%s", (player2_name,))
    player2_id = cur.fetchone()
    if player2_id is None:
      # If the player does not exist, insert them into the players table with a unique id
      cur.execute("SELECT nextval('player_id_seq')")
      id = cur.fetchone()[0]
      cur.execute("INSERT INTO player (player_id, first_name) VALUES (%s, %s)", (id, player2_name))
      player2_id = id
    else:
      # If the player already exists, retrieve their player_id
      player2_id = player2_id[0]
    # Print the first_name and player_id of the player
    print(f'Processing player2 {player2_name} with id {player2_id}')



  # Check if the player3 first_name is not empty
  if player3_name:
    # Check if the player already exists in the Players table
    cur.execute("SELECT player_id FROM player WHERE first_name=%s", (player3_name,))
    player3_id = cur.fetchone()
    if player3_id is None:
      # If the player does not exist, insert them into the players table with a unique id
      cur.execute("SELECT nextval('player_id_seq')")
      id = cur.fetchone()[0]
      cur.execute("INSERT INTO player (player_id, first_name) VALUES (%s, %s)", (id, player3_name))
      player3_id = id
    else:
      # If the player already exists, retrieve their player_id
      player3_id = player3_id[0]
    # Print the name and player_id of the player
    print(f'Processing player3 {player3_name} with id {player3_id}')


  # Check if the player4 first_name is not empty
  if player4_name:
    # Check if the player already exists in the Players table
    cur.execute("SELECT player_id FROM player WHERE first_name=%s", (player4_name,))
    player4_id = cur.fetchone()
    if player4_id is None:
      # If the player does not exist, insert them into the players table with a unique id
      cur.execute("SELECT nextval('player_id_seq')")
      id = cur.fetchone()[0]
      cur.execute("INSERT INTO player (player_id, first_name) VALUES (%s, %s)", (id, player4_name))
      player4_id = id
    else:
      # If the player already exists, retrieve their id
      player4_id = player4_id[0]
     # Print the name and id of the player
    print(f'Processing player4 {player4_name} with id {player4_id}')
  


  # Check if the team already exists in the Teams table
    cur.execute("SELECT team_id FROM team WHERE (team_player_1_id=%s AND team_player_2_id=%s) OR (team_player_1_id=%s AND team_player_2_id=%s)", (player1_id, player2_id, player2_id, player1_id))
    team_player_1_id = cur.fetchone()
    if team_player_1_id is None:
      # If the team does not exist, insert them into the teams table with a unique id
      cur.execute("SELECT nextval('team_id_seq')")
      id = cur.fetchone()[0]
      cur.execute("INSERT INTO team (team_id, team_player_1_id, team_player_2_id) VALUES (%s, %s, %s)", (id, player1_id, player2_id))
      team_player_1_id = id
    else:
      # If the team already exists, retrieve their id
      team_player_1_id = team_player_1_id[0]
    
    # Repeat the process for the other team
    cur.execute("SELECT team_id FROM team WHERE (team_player_1_id=%s AND team_player_2_id=%s) OR (team_player_1_id=%s AND team_player_2_id=%s)", (player3_id, player4_id, player4_id, player3_id))
    team_player_2_id = cur.fetchone()
    if team_player_2_id is None:
      cur.execute("SELECT nextval('team_id_seq')")
      id = cur.fetchone()[0]
      cur.execute("INSERT INTO team (team_id, team_player_1_id, team_player_2_id) VALUES (%s, %s, %s)", (id, player3_id, player4_id))
      team_player_2_id = id
    else:
      team_player_2_id = team_player_2_id[0]


  # Commit the changes to the database

  conn.commit()


# Close the cursor and connection
cur.close()
conn.close()