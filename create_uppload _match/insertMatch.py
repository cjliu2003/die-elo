import psycopg2
from openpyxl import load_workbook
from config import DATABASE_CONFIG



# Connect to the database
conn = psycopg2.connect(
    host=DATABASE_CONFIG['host'],
    database=DATABASE_CONFIG['database'],
    user=DATABASE_CONFIG['user'],
    password=DATABASE_CONFIG['password']
)

# Create a cursor
cur = conn.cursor()

# Load the workbook
wb = load_workbook('create_uppload _match/Fall 23 Die Tourney Data (1).xlsx')

# Select the active sheet
ws = wb.active

# Iterate through the rows of the sheet

for row in ws.iter_rows(min_row=3):

    # checking if the rows are not null
    if all(cell.value == None for cell in row):
        break
    date = row[0].value
    player1_name = row[1].value
    player2_name = row[2].value
    team1_score = row[3].value
    player3_name = row[4].value
    player4_name = row[5].value
    team2_score = row[6].value
    print(date, player1_name, player2_name, team1_score, player3_name, player4_name, team2_score)

    player_ids = []
    for i in range(1 , 6): 
      if (i == 3): 
        i += 1
      else:
          player_name = row[i].value
          # Check if the player name is not empty
          if player_name:
            # Check if the player already exists in the Player table
            cur.execute("SELECT player_id FROM player WHERE player_name=%s", (player_name,))
            player_id = cur.fetchone()
            if player_id is None:
              # If the player does not exist, insert them into the players table with a unique id
              cur.execute("SELECT nextval('player_id_seq')")
              id = cur.fetchone()[0]
              print(id, player_name)
              cur.execute("INSERT INTO player (player_id, player_name) VALUES (%s, %s)", (id, player_name))
              player_id = id
            else:
              # If the player already exists, retrieve their player_id
              player_id = player_id[0]
            player_ids.append(player_id)
    print("player ids: ", player_ids)

  # Check if the team already exists in the Teams table
    team_ids = []
    for i in range(2):
      if (i == 1):
          i += 1
      cur.execute("SELECT team_id FROM team WHERE (team_player_1_id=%s AND team_player_2_id=%s)", (player_ids[i], player_ids[i+1]))
      team_id = cur.fetchone()
      if team_id is None:
        # If the team does not exist, insert them into the teams table with a unique id
        cur.execute("SELECT nextval('team_id_seq')")
        team_id = cur.fetchone()[0]
        cur.execute("INSERT INTO team (team_id, team_player_1_id, team_player_2_id) VALUES (%s, %s, %s)", (team_id, player_ids[i], player_ids[i+1]))
      else:
        # If the team already exists, retrieve their id
        team_id = team_id[0]  
      team_ids.append(team_id)


  # Commit the changes to the database
    conn.commit()

    # Function to check winning team
    def get_winning_team(team1_score, team2_score):
      if team1_score > team2_score:
          return 1
      elif team2_score > team1_score:
          return 2
      else:
          return 0  


  # Insert the game into the matches table
    winning_team = get_winning_team(team1_score, team2_score)

    if winning_team == 1:
        winning_team_id = team_ids[0]
        losing_team_id = team_ids[1]
        winning_team_score = team1_score
        losing_team_score = team2_score
    elif winning_team == 2:
        winning_team_id = team_ids[1]
        losing_team_id = team_ids[0]
        winning_team_score = team2_score
        losing_team_score = team1_score
    else:
        winning_team_id = None
        losing_team_id = None
        winning_team_score = None
        losing_team_score = None  

    cur.execute("SELECT * FROM match WHERE match_timestamp=%s AND winning_team_id=%s AND losing_team_id=%s AND winning_team_score=%s AND losing_team_score=%s", (date, winning_team_id, losing_team_id, winning_team_score, losing_team_score))
    match = cur.fetchone()
    if match is None:
            cur.execute("INSERT INTO match (match_timestamp, winning_team_id, losing_team_id, winning_team_score, losing_team_score) VALUES (%s, %s, %s, %s, %s)", (date, winning_team_id, losing_team_id, winning_team_score, losing_team_score))
            print(f'processing match: {date} with {player1_name} and {player2_name} vs {player3_name} and {player4_name}: {team1_score} - {team2_score}  ')
            conn.commit()

            # get the last of the matches
            cur.execute("SELECT match_id FROM match ORDER BY match_id DESC LIMIT 1")
            match_id = cur.fetchone()[0]
          
            # Insert the players into the PlayerMatch table
            for i in range(4):
                cur.execute("INSERT INTO PlayerMatch (player_id,match_id) VALUES (%s, %s)", (player_ids[i], match_id))

            # Insert the team into the TeamMatch table
            cur.execute("INSERT INTO TeamMatch (team_id,match_id) VALUES (%s, %s)", (winning_team_id, match_id))
            cur.execute("INSERT INTO TeamMatch (team_id,match_id) VALUES (%s, %s)", (losing_team_id, match_id))
  
    else:
        print(f'Skipping match      : {date} with {player1_name} and {player2_name} vs {player3_name} and {player4_name}: {team1_score} - {team2_score}  , the match already exist')

    conn.commit()
     
conn.close()